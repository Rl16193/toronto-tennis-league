"""Export the converted sample dataset to one Excel workbook per collection.

    npm run dataset:xlsx
    py scripts/export-dataset-xlsx.py --out analysis/exports/new_database

Reads `tests/fixtures/dataset/` — the live snapshot already transformed into the post-remodel
shape by `scripts/build-sample-dataset.mjs` — and writes one .xlsx per collection so the new
database can be read, sorted and filtered without a Firestore client.

Each workbook has two sheets:

  data    one row per document. Columns follow the declared field order in
          `tests/fixtures/shape-reference.mjs` (via the generated `_schema.json`), so the sheet
          reads in the same order as the schema. Nested maps and arrays are written as JSON text;
          Firestore timestamps become ISO 8601 strings.
  schema  one row per field: how many documents carry it, how populated it is, a sample value,
          and whether the shape reference declares it. This is what makes the workbook
          self-documenting — a reader can see at a glance which fields are sparse.

Plus `_INDEX.xlsx`, listing every collection with its document count and the provenance of the
build (source snapshot, source project, whether names were pseudonymised).

The data is pseudonymised unless the dataset was built with --real-names; `_INDEX.xlsx` says
which, on the front sheet.
"""

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name=FONT, size=10)
NOTE_FONT = Font(name=FONT, size=10, italic=True, color="595959")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
RETIRED_FILL = PatternFill("solid", fgColor="FFF2CC")

# Excel refuses a cell value over 32,767 characters.
CELL_LIMIT = 32767


def to_cell(value):
    """Flatten one Firestore value into something a spreadsheet cell can hold."""
    if value is None:
        return ""
    if isinstance(value, bool):
        # Checked before int: bool is a subclass of int in Python.
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float, str)):
        text = value
    elif isinstance(value, dict) and "_seconds" in value:
        # A Firestore Timestamp as the JSON export leaves it.
        import datetime

        stamp = datetime.datetime.fromtimestamp(value["_seconds"], datetime.timezone.utc)
        return stamp.isoformat().replace("+00:00", "Z")
    else:
        text = json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(text, str) and len(text) > CELL_LIMIT:
        return text[: CELL_LIMIT - 20] + "…[truncated]"
    return text


def style_header(sheet, width_for):
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 22
    for index, width in enumerate(width_for, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def column_order(collection, docs, schema):
    """Declared fields first in schema order, then open-set fields, then anything undeclared.

    `tasks` stores each completed milestone tier as `true` under its catalogue id, so those column
    names are an open — but known — set. They are legitimate, and must not be counted as undeclared:
    a false warning on a valid column is worse than no warning at all.
    """
    declared = schema.get(collection, {}).get("fields", [])
    open_tiers = set(schema.get(collection, {}).get("openTierIds", []))
    present = set()
    for doc in docs:
        present.update(doc.get("data", {}).keys())
    ordered = [field for field in declared if field in present]
    tiers = sorted(present & open_tiers - set(ordered))
    extra = sorted(present - set(ordered) - set(tiers))
    return ordered + tiers, extra


def build_collection(collection, docs, schema, out_dir):
    ordered, extra = column_order(collection, docs, schema)
    fields = ordered + extra
    declared = set(schema.get(collection, {}).get("fields", []))
    open_tiers = set(schema.get(collection, {}).get("openTierIds", []))

    book = Workbook()
    data_sheet = book.active
    data_sheet.title = "data"

    headers = ["_path"] + fields
    data_sheet.append(headers)
    for doc in docs:
        row = [doc.get("path", "")]
        payload = doc.get("data", {})
        row.extend(to_cell(payload.get(field)) for field in fields)
        data_sheet.append(row)

    widths = [46] + [min(max(len(field) + 4, 12), 42) for field in fields]
    style_header(data_sheet, widths)
    for row in data_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT

    # --- schema sheet -------------------------------------------------------
    schema_sheet = book.create_sheet("schema")
    schema_sheet.append(["field", "documents_with_value", "populated_%", "declared_in_shape", "sample_value"])
    total = len(docs) or 1
    for field in fields:
        filled = sum(1 for doc in docs if doc.get("data", {}).get(field) not in (None, "", [], {}))
        sample = ""
        for doc in docs:
            value = doc.get("data", {}).get(field)
            if value not in (None, "", [], {}):
                sample = to_cell(value)
                break
        if isinstance(sample, str) and len(sample) > 200:
            sample = sample[:200] + "…"
        is_declared = field in declared or field in open_tiers
        schema_sheet.append(
            [
                field,
                filled,
                round(filled / total, 4),
                "yes" if is_declared else "NOT DECLARED",
                sample,
            ]
        )
        if not is_declared:
            for cell in schema_sheet[schema_sheet.max_row]:
                cell.fill = RETIRED_FILL
    style_header(schema_sheet, [34, 22, 14, 20, 60])
    for row in schema_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
        row[2].number_format = "0.0%"

    book.save(out_dir / f"{collection}.xlsx")
    return len(docs), len(fields), len(extra)


def build_index(manifest, summary, out_dir):
    book = Workbook()
    sheet = book.active
    sheet.title = "index"

    sheet["A1"] = "Converted database — new schema"
    sheet["A1"].font = TITLE_FONT
    provenance = [
        ("Built from", manifest.get("builtFrom", "")),
        ("Source project", f"{manifest.get('sourceProject', '')} (live deployed)"),
        ("Schema", manifest.get("schema", "")),
        (
            "Names",
            "pseudonymised — deterministic synthetic personas"
            if manifest.get("pseudonymised")
            else "REAL MEMBER DATA — do not share or publish",
        ),
        ("Total documents", manifest.get("totalDocs", "")),
    ]
    for offset, (label, value) in enumerate(provenance, start=3):
        sheet.cell(row=offset, column=1, value=label).font = Font(name=FONT, bold=True, size=10)
        sheet.cell(row=offset, column=2, value=value).font = BODY_FONT

    note_row = 3 + len(provenance) + 1
    sheet.cell(
        row=note_row,
        column=1,
        value=(
            "Each collection is one workbook: sheet 'data' holds the documents, sheet 'schema' shows how "
            "populated each field is. Nested maps and arrays are JSON text; timestamps are ISO 8601 (UTC). "
            "Documents are sparse — an empty cell means the field is absent, which is meaningful."
        ),
    ).font = NOTE_FONT
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    sheet.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[note_row].height = 46

    table_row = note_row + 2
    headers = ["collection", "documents", "columns", "undeclared_columns", "workbook"]
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=table_row, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for offset, row in enumerate(summary, start=table_row + 1):
        for index, value in enumerate(row, start=1):
            sheet.cell(row=offset, column=index, value=value).font = BODY_FONT

    for index, width in enumerate([32, 14, 12, 22, 30], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = sheet.cell(row=table_row + 1, column=1)

    book.save(out_dir / "_INDEX.xlsx")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dir", default="tests/fixtures/dataset", help="dataset directory to read")
    parser.add_argument("--out", default="analysis/exports/new_database", help="directory to write workbooks into")
    args = parser.parse_args()

    dataset_dir = (ROOT / args.dir).resolve()
    out_dir = (ROOT / args.out).resolve()

    manifest_path = dataset_dir / "_manifest.json"
    schema_path = dataset_dir / "_schema.json"
    if not manifest_path.exists():
        sys.exit(f"No dataset at {dataset_dir}. Build it first: npm run dataset:build")
    if not schema_path.exists():
        sys.exit(f"No _schema.json in {dataset_dir}. Rebuild the dataset: npm run dataset:build")

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    out_dir.mkdir(parents=True, exist_ok=True)

    summary = []
    for path in sorted(dataset_dir.glob("*.json")):
        if path.name.startswith("_"):
            continue
        collection = path.stem
        docs = json.loads(path.read_text(encoding="utf-8"))
        count, columns, undeclared = build_collection(collection, docs, schema, out_dir)
        summary.append([collection, count, columns, undeclared, f"{collection}.xlsx"])

    build_index(manifest, summary, out_dir)

    print(f"Wrote {len(summary)} workbooks + _INDEX.xlsx to {out_dir.relative_to(ROOT)}")
    print(f"  source        {manifest.get('builtFrom')} ({manifest.get('sourceProject')})")
    print(f"  pseudonymised {manifest.get('pseudonymised')}")
    print()
    for collection, count, columns, undeclared, _ in summary:
        flag = f"   {undeclared} undeclared column(s)" if undeclared else ""
        print(f"  {collection:<26} {count:>5} rows  {columns:>3} cols{flag}")


if __name__ == "__main__":
    main()
